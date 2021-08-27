from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#create a workbook from objects
#wb = Workbook()


#load exiting worksheet
wb = load_workbook('data.xlsx')

# create a active worksheet
ws = wb.active

#set a variable
name = ws['A2'].value
color = ws['B3'].value
column_a = ws['A']
row_1 = ws['1']
range = ws['A2':'A10']
range2 = ws['A3':'A4']
#print(range)

for cell in range:
    for x in cell:
        print(x.value)
#print something from spreadsheet
#print(f'{name} {color}')