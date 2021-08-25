from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('Computer')
root.geometry('500x800')
# Create workbook instance
wb = Workbook()
#Load workbook 
wb = load_workbook('data.xlsx')
# Create active worsheet
ws =wb.active
column_a =ws['A']
column_b = ws['B']

def get_a():
    list =""
    for cell in column_a:
        list = f'{list + str(cell.value)}\n'
    label_a.config(text=list)
    
def get_b():
    list =""
    for cell in column_b:
        list = f'{list + str(cell.value)}\n'
    label_b.config(text=list)    

#create button to show data
ba = Button(root, text = "Get Column A",command=get_a)
ba.pack(pady=20)
#show data from button
label_a  = Label(root, text='')
label_a.pack(pady=20)


bb = Button(root, text = "Get Column B",command=get_b)
bb.pack(pady=20)
label_b  = Label(root, text='')
label_b.pack(pady=20)

#add new data in spreadsheet
ws['A8'] = "Fred"
ws['B8'] = "Cheese"
# save into new file
wb.save('newdata.xlsx')
root.mainloop()